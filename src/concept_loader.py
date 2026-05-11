"""
概念类目树加载模块
"""
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field


@dataclass
class ConceptNode:
    """概念节点"""
    name: str
    level: int  # 1, 2, 3, or 4 (leaf)
    children: List['ConceptNode'] = field(default_factory=list)
    parent: Optional['ConceptNode'] = None
    path: List[str] = field(default_factory=list)  # 从根到当前节点的路径
    
    def is_leaf(self) -> bool:
        return len(self.children) == 0
    
    def get_level3_category(self) -> Optional[str]:
        """获取所属的三级类目名称"""
        if self.level == 3:
            return self.name
        elif self.level == 4 and len(self.path) >= 3:
            return self.path[2]
        return None


@dataclass
class CategorySheet:
    """类目表结构"""
    name: str
    level1_categories: List[ConceptNode] = field(default_factory=list)
    level3_categories: List[ConceptNode] = field(default_factory=list)  # 所有三级类目
    leaves: List[ConceptNode] = field(default_factory=list)  # 所有叶子节点


class ConceptLoader:
    """概念类目树加载器"""
    
    SHEET_NAMES = ['主体', '运动', '场景', '音频类型']
    
    def __init__(self, xlsx_path: str, sheet_names: Optional[List[str]] = None):
        self.xlsx_path = Path(xlsx_path)
        self.sheet_names = sheet_names or list(self.SHEET_NAMES)
        self.workbook = None
        self.categories: Dict[str, CategorySheet] = {}
    
    def load(self) -> Dict[str, CategorySheet]:
        """加载所有概念类目"""
        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"概念类目树文件不存在: {self.xlsx_path}")
        
        self.workbook = openpyxl.load_workbook(self.xlsx_path, read_only=True)
        
        for sheet_name in self.sheet_names:
            if sheet_name in self.workbook.sheetnames:
                self.categories[sheet_name] = self._parse_sheet(sheet_name)
        
        self.workbook.close()
        return self.categories
    
    def _parse_sheet(self, sheet_name: str) -> CategorySheet:
        """解析单个sheet"""
        sheet = self.workbook[sheet_name]
        category = CategorySheet(name=sheet_name)
        
        # 遍历每一行，构建树结构
        current_level1: Optional[ConceptNode] = None
        current_level2: Optional[ConceptNode] = None
        current_level3: Optional[ConceptNode] = None
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):
                continue
            
            # 安全地解析行数据
            row_list = list(row) if row else []
            level1_val = row_list[0] if len(row_list) > 0 else None
            level2_val = row_list[1] if len(row_list) > 1 else None
            level3_val = row_list[2] if len(row_list) > 2 else None
            leaf_val = row_list[3] if len(row_list) > 3 else None
            
            # 一级类目
            if level1_val is not None:
                current_level1 = ConceptNode(
                    name=str(level1_val),
                    level=1,
                    path=[str(level1_val)]
                )
                category.level1_categories.append(current_level1)
                current_level2 = None
                current_level3 = None
            
            # 二级类目
            if level2_val is not None and current_level1 is not None:
                current_level2 = ConceptNode(
                    name=str(level2_val),
                    level=2,
                    parent=current_level1,
                    path=current_level1.path + [str(level2_val)]
                )
                current_level1.children.append(current_level2)
                current_level3 = None
            
            # 三级类目
            if level3_val is not None:
                parent = current_level2 if current_level2 else current_level1
                if parent:
                    current_level3 = ConceptNode(
                        name=str(level3_val),
                        level=3,
                        parent=parent,
                        path=parent.path + [str(level3_val)]
                    )
                    parent.children.append(current_level3)
                    category.level3_categories.append(current_level3)
            
            # 叶片节点
            if leaf_val is not None and current_level3 is not None:
                # 叶片可能是逗号分隔的多个值
                leaf_names = [l.strip() for l in str(leaf_val).replace('，', ',').split(',') if l.strip()]
                for leaf_name in leaf_names:
                    leaf_node = ConceptNode(
                        name=leaf_name,
                        level=4,
                        parent=current_level3,
                        path=current_level3.path + [leaf_name]
                    )
                    current_level3.children.append(leaf_node)
                    category.leaves.append(leaf_node)
        
        return category
    
    def get_level3_categories(self, sheet_name: str) -> List[ConceptNode]:
        """获取指定sheet的所有三级类目"""
        if sheet_name not in self.categories:
            return []
        return self.categories[sheet_name].level3_categories
    
    def get_leaves_under_level3(self, sheet_name: str, level3_name: str) -> List[ConceptNode]:
        """获取指定三级类目下的所有叶片"""
        if sheet_name not in self.categories:
            return []
        
        for node in self.categories[sheet_name].level3_categories:
            if node.name == level3_name:
                return node.children
        return []
    
    def get_all_level3_with_leaves(self) -> Dict[str, Dict[str, List[str]]]:
        """获取所有三级类目及其叶片（用于统计）"""
        result = {}
        for sheet_name, category in self.categories.items():
            result[sheet_name] = {}
            for level3 in category.level3_categories:
                leaves = [leaf.name for leaf in level3.children]
                result[sheet_name][level3.name] = leaves
        return result
    
    def get_stats(self) -> Dict[str, Any]:
        """获取统计信息"""
        stats = {}
        for sheet_name, category in self.categories.items():
            stats[sheet_name] = {
                'level1_count': len(category.level1_categories),
                'level3_count': len(category.level3_categories),
                'leaf_count': len(category.leaves)
            }
        return stats


def load_concepts(xlsx_path: str, sheet_names: Optional[List[str]] = None) -> ConceptLoader:
    """加载概念类目树的便捷函数"""
    loader = ConceptLoader(xlsx_path, sheet_names=sheet_names)
    loader.load()
    return loader
